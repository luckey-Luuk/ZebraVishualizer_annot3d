from PySide6.QtCore import Qt#, QSize
from PySide6.QtGui import QColor, QIcon, QKeySequence, QPalette, QPixmap, QColor, QAction#, QResizeEvent # move QAction to QtWidgets when using Python 3.10.10
from PySide6.QtWidgets import QApplication, QComboBox, QDockWidget, QFileDialog, QDialog, QHBoxLayout, QInputDialog, QLabel, QMainWindow, QPushButton, QSlider, QSpinBox, QGridLayout, QVBoxLayout, QWidget, QTabWidget, QFrame#, QAction

# from traits.etsconfig.api import ETSConfig
# ETSConfig.toolkit = 'qt4' # fix traitsui.qt4.* modules having moved to traitsui.qt.*

from traits.api import HasTraits, Instance, on_trait_change
from traitsui.api import View, Item
from mayavi.core.ui.api import MayaviScene, MlabSceneModel, SceneEditor
from mayavi import mlab

import numpy as np
import sys
import math
import pickle
from openpyxl import Workbook, load_workbook

from AnnotationSpace3D import AnnotationSpace3D
from helpers import find_centroids, read_tiff, create_image_dict, create_colour_array

#TODO: get rid of globals

COLORS = {
    '#ff0000': [255, 0, 0, 255],
    '#35e3e3': [53, 227, 227, 255],
    '#5ebb49': [94, 187, 73, 255],
    '#ffd035': [255, 208, 53, 255],
}

# ERASER_COLOR_RGBA = [255, 255, 255, 255]
INIT_COLOR_RGBA = COLORS['#ff0000']

# p = 'xy' # xy initially, yz, xz TODO: kijken of hier iets mee wordt gedaan
current_frame = {'xy': 0, 'xz': 0, 'yz': 0}
annot3D = -1
w, h, d = 500, 500, 25

amount_of_points = 20
# directory = 'data/data_annot3d/20190621++2' #folder containing tif files
directory = 'data/data_ZebraVishualizer/original/3D tracking data to visualize/20190701--2_inter_29layers_mask_3a'


def get_filled_pixmap(pixmap_file):
    pixmap = QPixmap(pixmap_file)
    mask = pixmap.createMaskFromColor(QColor('black'), Qt.MaskOutColor)
    pixmap.fill((QColor('white')))
    pixmap.setMask(mask)
    return pixmap


class Visualization(HasTraits):
    scene = Instance(MlabSceneModel, ())

    @on_trait_change('scene.activated')
    

    def update_plot(self): #initializatie
        global directory
        self.image_dictionary = create_image_dict(directory)
        self.amount_of_frames = len(self.image_dictionary)
        self.current_frame_number = 0
        self.current_point_index = 0
        self.transparancy = 1.0

        self.show_volume = True
        self.show_trajectory = False
        self.show_all_trajectories = False

        self.colour_array = create_colour_array()
        
        self.point_location_data=[{} for f in range(self.amount_of_frames)] #information of point locations in every time step
        
        self.mayavi_trajectory_dots = [{} for f in range(self.amount_of_frames)] #for mayvi to store points for the display view

        self.mayavi_trajectory_lines = [{} for f in range(self.amount_of_frames)] #for mayavi to store lines for the display view

        self.mayavi_dots = {}#location for mayavi to store individual dots

        self.figure = mlab.gcf(engine=self.scene.engine)#nodig voor de picker functie

        # global annot3D
        npimages = annot3D.get_npimages()

        self.x_lenght = len(npimages[0][0])
        self.y_lenght = len(npimages[0])
        self.z_lenght = len(npimages)

        self.sphere_size = 10

        npspace = annot3D.get_npspace()
        self.npspace_sf = mlab.pipeline.scalar_field(npspace) # scalar field to update later
        self.volume = mlab.pipeline.volume(mlab.pipeline.scalar_field(npimages),color=(0,1,0),vmin=0,vmax=np.amax(npimages)*self.transparancy) #render het volume voor de eerste keer

        self.figure.on_mouse_pick(self.picker_callback) #zet picker aan

        segmask = mlab.pipeline.iso_surface(self.npspace_sf, color=(1.0, 0.0, 0.0))
        self.scene.background = (0.1, 0.1, 0.1)  
        # mlab.orientation_axes()

    def draw_point(self,new_x,new_y,new_z):#updates point data and draws updated point
        #update point data
        self.point_location_data[self.current_frame_number][self.current_point_index]=[new_x,new_y,new_z]
        #draw new point
        if self.current_point_index in self.mayavi_dots:
            self.mayavi_dots[self.current_point_index].remove()
            del self.mayavi_dots[self.current_point_index]
            
        self.mayavi_dots[self.current_point_index] = mlab.points3d(new_x,new_y,new_z,color=self.colour_array[self.current_point_index%len(self.colour_array)],scale_factor=self.sphere_size)

    def draw_previous_point(self): #places the point in the same location as it was in the previous image number
        if self.current_frame_number!=0 and self.current_point_index in self.point_location_data[self.current_frame_number-1]:
            new_location = self.point_location_data[self.current_frame_number-1][self.current_point_index]
            self.draw_point(new_location[0],new_location[1],new_location[2])

    def delete_point(self):
        if self.current_point_index in self.point_location_data[self.current_frame_number]:
            del self.point_location_data[self.current_frame_number][self.current_point_index]
        if self.current_point_index in self.mayavi_dots:
            self.mayavi_dots[self.current_point_index].remove()
            del self.mayavi_dots[self.current_point_index]

    def delete_all_points(self): #TODO: kijk of dit weg/samengevoegd met delete_point kan
        for p in range(amount_of_points):
            if p in self.mayavi_dots:
                self.mayavi_dots[p].remove()
                del self.mayavi_dots[p]

    def redraw_all_points(self): #redraws all points, used to update points for the next timestep
        self.delete_all_points()
        for p in range(amount_of_points):
            if p in self.point_location_data[self.current_frame_number]:
                self.mayavi_dots[p] = mlab.points3d(self.point_location_data[self.current_frame_number][p][0],self.point_location_data[self.current_frame_number][p][1],self.point_location_data[self.current_frame_number][p][2],color = self.colour_array[p%len(self.colour_array)],scale_factor=self.sphere_size)

    def add_value_to_point(self,added_value):
        if self.current_point_index in self.point_location_data[self.current_frame_number]:
            old_value = self.point_location_data[self.current_frame_number][self.current_point_index]
            self.draw_point(old_value[0]+added_value[0],old_value[1]+added_value[1],old_value[2]+added_value[2])

    def update_volume(self):
        window.load_source_file(directory+'/'+self.image_dictionary[self.current_frame_number])
        npimages = annot3D.get_npimages()
        if self.volume is not None:
            self.volume.remove()
            self.npspace_sf.remove()
        npspace = annot3D.get_npspace()
        self.npspace_sf = mlab.pipeline.scalar_field(npspace)
        self.volume = mlab.pipeline.volume(mlab.pipeline.scalar_field(npimages),color=(0,1,0),vmax=np.amax(npimages)*self.transparancy) #volume renderen
    
    def remove_volume(self):
        if self.volume is not None: #can't be removed if it isn't there in the first place
            self.volume.remove()
            self.volume = None
            self.npspace_sf.remove()
            self.npspace_sf = None
    
    def toggle_volume(self): #toggle voor volume (knop), gebruikt update en remove volume
        if self.show_volume==True:
            self.show_volume = False
            self.remove_volume()
        else:
            self.show_volume = True
            self.update_volume()
    
    def draw_trajectory_step(self, frame, point_index): #tekent trajectory #TODO: teken 1 stap
        if frame!=0 and point_index in self.point_location_data[frame]: #check if point exists #teken punt
            x_coordinate = self.point_location_data[frame][point_index][0]
            y_coordinate = self.point_location_data[frame][point_index][1]
            z_coordinate = self.point_location_data[frame][point_index][2]
            self.mayavi_trajectory_dots[frame][point_index] = mlab.points3d(x_coordinate,y_coordinate,z_coordinate,color=self.colour_array[point_index%len(self.colour_array)],scale_factor=3)

            if point_index in self.point_location_data[frame-1]: #check if previous point is not None #teken buis ertussen als er twee punten achter elkaar zijn
                x_coordinates = [self.point_location_data[frame-1][point_index][0],x_coordinate]
                y_coordinates = [self.point_location_data[frame-1][point_index][1],y_coordinate]
                z_coordinates = [self.point_location_data[frame-1][point_index][2],z_coordinate]
                self.mayavi_trajectory_lines[frame][point_index] = mlab.plot3d(x_coordinates,y_coordinates,z_coordinates,color=self.colour_array[point_index%len(self.colour_array)],tube_radius=1) # gebruik color=(0,0.9,0) voor groen
    
    def draw_trajectory(self, frame, point_index):
        for f in range(frame+1):
            self.draw_trajectory_step(f, point_index)

    def remove_trajectory_step(self, f, p):
        if p in self.mayavi_trajectory_dots[f]:
            self.mayavi_trajectory_dots[f][p].remove()
            del self.mayavi_trajectory_dots[f][p]
        
        if p in self.mayavi_trajectory_lines[f]:
            self.mayavi_trajectory_lines[f][p].remove()
            del self.mayavi_trajectory_lines[f][p]
    
    def remove_trajectory(self, frame, point_index):
        for f in range(frame+1):
            self.remove_trajectory_step(f, point_index)

    def remove_all_trajectories(self): #stop met trajectory visualiseren #TODO: verwijder 1 stap
        for p in range(amount_of_points-1):
            self.remove_trajectory(self.amount_of_frames-1, p)
        

    def toggle_trajectory(self): #changes showing and not showing results #toggle voor trajectory (knop)
        if self.show_trajectory==False:
            self.show_trajectory = True
            self.draw_trajectory(self.current_frame_number, self.current_point_index)
        else:
            self.remove_all_trajectories()
            self.show_trajectory = False

    def toggle_all_trajectories(self):
        if self.show_all_trajectories==False:
            self.show_all_trajectories = True
            for p in range(amount_of_points):
                self.draw_trajectory(self.current_frame_number, p)
        else:
            self.remove_all_trajectories()
            self.show_all_trajectories = False

    #gesplitst in next_frame, previous_frame en goto_frame
    # def update_frame(self,next_or_previous='next'):
    #     if next_or_previous=="next": #go to next frame
    #         if self.current_frame_number==self.amount_of_frames-1:
    #             self.current_frame_number = 0 #loop around
    #         else:
    #             self.current_frame_number += 1
    #     elif next_or_previous=="previous": #go to previous frame
    #         if self.current_frame_number==0:
    #             self.current_frame_number = self.amount_of_frames-1 #loop around
    #         else:
    #             self.current_frame_number -= 1
    #     elif isinstance(next_or_previous,int): #used for goto function to go to a specific frame
    #         if self.amount_of_frames-1<next_or_previous:
    #             self.current_frame_number = self.amount_of_frames-1
    #         else:
    #             self.current_frame_number = next_or_previous

    #     if self.show_volume==True:
    #         self.update_volume()

    #     self.redraw_all_points()
        
    #     self.remove_all_trajectories()
    #     if self.show_all_trajectories==True:
    #         for p in range(amount_of_points):
    #             self.draw_trajectory(self.current_frame_number, p)
    #     elif self.show_trajectory==True:
    #         self.draw_trajectory(self.current_frame_number, self.current_point_index)
    #     #mlab.orientation_axes()

    def next_frame(self):
        if self.current_frame_number==self.amount_of_frames-1:
            self.current_frame_number = 0 #loop around

            self.remove_all_trajectories()
            if self.show_all_trajectories==True:
                for p in range(amount_of_points):
                    if p in self.point_location_data[self.current_frame_number]:
                        self.draw_trajectory(self.current_frame_number, p)
            elif self.show_trajectory==True:
                if self.current_point_index in self.point_location_data[self.current_frame_number]:
                    self.draw_trajectory(self.current_frame_number, self.current_point_index)

        else:
            self.current_frame_number += 1

            if self.show_all_trajectories==True:
                for p in range(amount_of_points):
                    if p in self.point_location_data[self.current_frame_number]:
                        if p in self.point_location_data[self.current_frame_number-1]:
                            self.draw_trajectory_step(self.current_frame_number, p)
                        else:
                            self.draw_trajectory(self.current_frame_number, p)
                    else:
                        # for f in range(self.current_frame_number+1,self.amount_of_frames): #TODO: alleen trajectory niet laten zien als die niet meer komt
                        #     if p in self.point_location_data[f]:
                        self.remove_trajectory(self.current_frame_number, p)
            elif self.show_trajectory==True:
                if self.current_point_index in self.point_location_data[self.current_frame_number]:
                    if self.current_point_index in self.point_location_data[self.current_frame_number-1]:
                        self.draw_trajectory_step(self.current_frame_number, self.current_point_index)
                    else:
                        self.draw_trajectory(self.current_frame_number, self.current_point_index)
                else:
                    self.remove_trajectory(self.current_frame_number, self.current_point_index)

        if self.show_volume==True:
            self.update_volume()

        self.redraw_all_points()
        #mlab.orientation_axes()

    def previous_frame(self):
        if self.current_frame_number==0:
            self.current_frame_number = self.amount_of_frames-1 #loop around

            self.remove_all_trajectories()
            if self.show_all_trajectories==True:
                for p in range(amount_of_points-1):
                    if p in self.point_location_data[self.current_frame_number]:
                        self.draw_trajectory(self.current_frame_number, p)
            elif self.show_trajectory==True:
                if self.current_point_index in self.point_location_data[self.current_frame_number]:
                    self.draw_trajectory(self.current_frame_number, self.current_point_index)

        else:
            self.current_frame_number -= 1

            if self.show_all_trajectories==True:
                for p in range(amount_of_points-1):
                    if p in self.point_location_data[self.current_frame_number+1]:
                        if p in self.point_location_data[self.current_frame_number]:
                            self.remove_trajectory_step(self.current_frame_number+1, p)
                        else:
                            self.remove_trajectory(self.current_frame_number+1, p)
                    elif p in self.point_location_data[self.current_frame_number]:
                        self.draw_trajectory(self.current_frame_number, p)
            elif self.show_trajectory==True:
                if self.current_point_index in self.point_location_data[self.current_frame_number+1]:
                    if self.current_point_index in self.point_location_data[self.current_frame_number]:
                        self.remove_trajectory_step(self.current_frame_number+1, self.current_point_index)
                    else:
                        self.remove_trajectory(self.current_frame_number+1, self.current_point_index)
                elif self.current_point_index in self.point_location_data[self.current_frame_number]:
                    self.draw_trajectory(self.current_frame_number, self.current_point_index)

        if self.show_volume==True:
            self.update_volume()

        self.redraw_all_points()

        #mlab.orientation_axes()

    def goto_frame(self, frame_number):
        if self.amount_of_frames-1<frame_number:
            self.current_frame_number = self.amount_of_frames-1
        else:
            self.current_frame_number = frame_number

        if self.show_volume==True:
            self.update_volume()

        self.redraw_all_points()
        
        self.remove_all_trajectories()
        if self.show_all_trajectories==True:
            for p in range(amount_of_points-1):
                if p in self.point_location_data[self.current_frame_number]:
                    self.draw_trajectory(self.current_frame_number, p)
        elif self.show_trajectory==True:
            if self.current_point_index in self.point_location_data[self.current_frame_number]:
                self.draw_trajectory(self.current_frame_number, self.current_point_index)
        #mlab.orientation_axes()

    def picker_callback(self,picker): #kijkt waar je klikt en tekent punt
        if self.show_volume==True:
            #print(dir(picker))
            coordinates = picker.pick_position
            self.draw_point(coordinates[0],coordinates[1],coordinates[2])

    def save_trajectory(self,file_name = "test_file"): #knop save annotations, sla lijst met punten op in excel
        point_data_list = [] #meant to put in data from point_location_data that is not None, later used for export to excel
        for f in range(self.amount_of_frames):
            for p in range(amount_of_points):
                if p in self.point_location_data[f]:
                    point_data_list.append([f,p,self.point_location_data[f][p][0],self.point_location_data[f][p][1],self.point_location_data[f][p][2]])

        book = Workbook()
        sheet = book.active
        sheet.append(["timestep","dot","x","y","z"])
        for row in point_data_list:
            sheet.append(row)
        book.save(file_name)

    def export_trajectory(self,file_name,new_x_size,new_y_size,new_z_size): #zelfde als save data, maar met meer: rekent afstand tussen punten uit en voer werkelijke grootte in
        x_mod = new_x_size/self.x_lenght #the multiplyer to correct the cordinates to the real size
        y_mod = new_y_size/self.y_lenght
        z_mod = new_z_size/self.z_lenght
        point_data_list = [] #meant to put in data from point_location_data that is not None, later used for export to excel
        for f in range(self.amount_of_frames): #for every slice number
            for p in range(amount_of_points): #for every tracking number
                if p in self.point_location_data[f][p][0]:
                    x = self.point_location_data[f][p][0]*x_mod
                    y = self.point_location_data[f][p][1]*y_mod
                    z = self.point_location_data[f][p][2]*z_mod
                    point_data_list.append([p,f,z,y,x,-1]) #the x and z are switched during data import, that is why they are switched back in export
        encounterd_points = []
        for i in range(len(point_data_list)):
            for j in encounterd_points:
                if point_data_list[i][0]==j[0] and point_data_list[i][1]==j[1]+1:
                    encounterd_points.remove(j)
                    x_exponent = math.pow(point_data_list[i][2]-j[2],2)
                    y_exponent = math.pow(point_data_list[i][3]-j[3],2)
                    z_exponent = math.pow(point_data_list[i][4]-j[4],2)
                    point_data_list[i][5] = math.sqrt(x_exponent+y_exponent+z_exponent)
                    break
            encounterd_points.append(point_data_list[i])

        book = Workbook()
        sheet = book.active
        sheet.append(["tracking number","slice number","x","y","z","Distance"])
        for row in point_data_list:
            sheet.append(row)
        book.save(file_name)

    def load_xlsx(self,file_name="test_file.xlsx"): #knop load annotations TODO: change self.amount_of_points to max value in dot column
        book = load_workbook(filename=file_name)
        sheet = book.active

        global amount_of_points
        amount_of_points = max(cell.value for cell in sheet['B'][1:])+1
        self.point_location_data = [{} for f in range(self.amount_of_frames)] #set data back to None to remove old data
        self.mayavi_dots = {}

        for row in sheet.values:
            if type(row[0])==int: #done to skip the first row that doesn't give data
                self.point_location_data[row[0]][row[1]] = [row[2],row[3],row[4]]
        self.redraw_all_points()
    
    def load_pkl(self, pkl_dict):
        centroids = {}
        for f in range(self.amount_of_frames):
            centroids[f] = find_centroids(directory+'/'+self.image_dictionary[f])

        linked_centroids = {}
        for trajectory in pkl_dict:
            new_link = {}
            for link in trajectory:
                (curr_label, f, prev_label) = link
                new_link[f] = centroids[f][curr_label]
            linked_centroids[pkl_dict.index(trajectory)] = new_link
        
        global amount_of_points
        amount_of_points = len(pkl_dict)
        self.point_location_data = [{} for f in range(self.amount_of_frames)]
        self.mayavi_dots = {}
        for f in range(self.amount_of_frames):
            for p in range(amount_of_points):
                if f in linked_centroids[p]:
                    self.point_location_data[f][p][0] = linked_centroids[p][f][0]
                    self.point_location_data[f][p][1] = linked_centroids[p][f][1]
                    self.point_location_data[f][p][2] = linked_centroids[p][f][2]

    def update_annot(self): # update the scalar field and visualization auto updates
        npspace = annot3D.get_npspace()
        self.npspace_sf.mlab_source.trait_set(scalars=npspace)
    
    view = View(Item('scene', editor=SceneEditor(scene_class=MayaviScene), height=250, width=300, show_label=False), resizable=True )

class MayaviQWidget(QWidget): #mayavi raam
    def __init__(self, parent=None):
        QWidget.__init__(self, parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)
        layout.setSpacing(0)
        self.visualization = Visualization()

        self.ui = self.visualization.edit_traits(parent=self, kind='subpanel').control
        layout.addWidget(self.ui)
        self.ui.setParent(self)
    
    def update_annot(self):
        self.visualization.update_annot()

class MainWindow(QMainWindow): #hele raam
    c = {'xy': 0, 'xz': 0, 'yz': 0}
    
    # dims = (500, 500, 25) # w, h, d TODO: kijken of hier iets mee wordt gedaan
    frames = {}
    # plane_depth = {} TODO: kijken of hier iets mee wordt gedaan
    # frame_annotations = {} TODO: kijken of hier iets mee wordt gedaan
    # num_frames = 0 TODO: kijken of hier iets mee wordt gedaan
    npimages = -1

    pkl_dict = None


    def load_tiff_dialog(self):
        dname = QFileDialog.getExistingDirectory(self, 'Select folder containing .tif files')

        if dname:
            global directory
            directory = dname
        # else:
        #     sys.exit()

    def load_annot_dialog(self, setup=False): #TODO: check if .pkl is compatible with .tif
        fname, _ = QFileDialog.getOpenFileName(self, 'Select file containing trajectories (optional)', '.', filter="(*.pkl *.xlsx)")

        if fname.endswith('.pkl'):
            key = directory.split('/')[-1].split('_')[0]
            self.pkl_dict = pickle.load(open(fname, 'rb'))[key]
            self.mayavi_widget.visualization.load_pkl(self.pkl_dict)
        elif fname.endswith('.xlsx'):
            self.mayavi_widget.visualization.load_xlsx(fname)
        else:
            return

        if not setup:
            window.mayavi_widget.visualization.goto_frame(self.mayavi_widget.visualization.current_frame_number)
            # window.mayavi_widget.visualization.redraw_all_points()
            # if window.mayavi_widget.visualization.show_trajectory==True:
            #     window.mayavi_widget.visualization.draw_trajectory(self.mayavi_widget.visualization.current_frame_number, self.mayavi_widget.visualization.current_point_index)
        return

    def load_source_file(self, filename): #uit annot3D, nodig
        """Text.

        Parameters
        ----------
        filename : ???
            explanation

        Returns
        -------
        None
        """
        # global COLORS, p, current_frame, annot3D
        global annot3D

        self.frames['xy'], self.frames['xz'], self.frames['yz'] = read_tiff(filename)

        self.npimages = self.frames['xy']
        
        w = self.npimages[0].shape[0]
        h = self.npimages[0].shape[1]
        d = self.npimages.shape[0]

        annot3D = AnnotationSpace3D(self.npimages, (d, w, h), INIT_COLOR_RGBA)

        # self.plane_depth = {
        #     'xy': d,
        #     'xz': w,
        #     'yz': h
        # }

        # self.dims = (w, h, d)

        # self.num_frames = self.plane_depth[p]

        # self.frame_annotations = {
        #     'xy': [] * d, 
        #     'xz': [] * w, 
        #     'yz': [] * h
        # }
    
    def animate(self):
        @mlab.animate(delay=100)
        def anim():
            while True:
                self.change_volume_model_next()
                yield
        anim()

    def __init__(self):
        """Text.

        Parameters
        ----------
        filename : ???
            explanation

        Returns
        -------
        None
        """
        super().__init__()
        
    # INIT ANNOT LOAD UP
        self.load_tiff_dialog()
        temp_dict = create_image_dict(directory)
        self.load_source_file(directory + '/' + temp_dict[0])

        main_widget = QWidget()
        main_layout = QVBoxLayout()
        main_widget.setLayout(main_layout)
        main_widget.setMaximumWidth(300)
        self.setWindowTitle("ZebraVishualizer_annot3d")
        self.setCentralWidget(main_widget)

        tab = QTabWidget(self)
        options_page = QWidget()
        cell_page = QWidget()


    # MAYAVI RENDER VIEW
        self.rdock = QDockWidget("Render View", self) # render dock
        self.rdock.setFeatures(self.rdock.features() & ~QDockWidget.DockWidgetClosable) # unclosable
        self.rdock.setAllowedAreas(Qt.LeftDockWidgetArea | Qt.RightDockWidgetArea)
        container = QWidget(self.rdock)
        self.mayavi_widget = MayaviQWidget(container)
        self.rdock.setWidget(self.mayavi_widget)
        self.addDockWidget(Qt.RightDockWidgetArea, self.rdock)

        self.load_annot_dialog(True)

        self.animate()


    # OPTIONS PAGE
        options_page_layout = QVBoxLayout()
        options_page.setLayout(options_page_layout)

        annotation_layout = QVBoxLayout()
        visualization_layout = QVBoxLayout()

        def create_header(text, layout):
            label = QLabel(text)
            layout.addWidget(label)

            line = QFrame();
            line.setFrameShape(QFrame.HLine);
            line.setFrameShadow(QFrame.Sunken);
            layout.addWidget(line)

        # ANNOTATION OPTIONS LAYOUT
        create_header('Annotation options', annotation_layout)

        annotation_buttons_layout = QGridLayout()

        def change_selected_point(new_point): #leest welke aangeklikt is
            new_point = new_point.split(" ")[1] #split the string and take the number
            self.mayavi_widget.visualization.current_point_index = int(new_point)
            if self.mayavi_widget.visualization.show_trajectory==True: #change between different results if mode is result
                self.mayavi_widget.visualization.remove_all_trajectories()
                self.mayavi_widget.visualization.draw_trajectory(self.mayavi_widget.visualization.current_frame_number, self.mayavi_widget.visualization.current_point_index)

        point_list = []   #list for the selectable points in the combobox
        for i in range(amount_of_points): #TODO: verander als trajectory met meer punten wordt geladen
            point_list.append("cell " + str(i))

        selection_box = QComboBox()
        selection_box.addItems(point_list)    
        selection_box.currentIndexChanged.connect(lambda: change_selected_point(selection_box.currentText()))
        annotation_buttons_layout.addWidget(selection_box, 0, 0)

        self.continue_button = QPushButton('copy previous\nannotation')
        self.continue_button.clicked.connect(lambda: self.mayavi_widget.visualization.draw_previous_point())
        annotation_buttons_layout.addWidget(self.continue_button, 0, 1)

        self.delete_button = QPushButton('delete\nannotation') #'delete' is wat er op knop staat
        self.delete_button.clicked.connect(lambda: self.mayavi_widget.visualization.delete_point()) #connectie tussen knop en functie
        annotation_buttons_layout.addWidget(self.delete_button, 0, 2) #plaats knop op grid (self.knop, x op grid, y op grid)

        annotation_layout.addLayout(annotation_buttons_layout)


        # Adjustments layout
        adjustments_layout = QGridLayout()

        adjustments_layout.setRowMinimumHeight(0, 10)
        adjustments_layout.setColumnMinimumWidth(0, 45)
        
        x_label = QLabel('X')
        x_label.setFixedWidth(25)
        adjustments_layout.addWidget(x_label, 1, 1)
        y_label = QLabel('Y')
        y_label.setFixedWidth(25)
        adjustments_layout.addWidget(y_label, 2, 1)
        z_label = QLabel('Z')
        z_label.setFixedWidth(25)
        adjustments_layout.addWidget(z_label, 3, 1)

        def create_adjust_button(self, text, update=[0,0,0]):
            self.button = QPushButton(text)
            self.button.setFixedSize(25, 25)
            self.button.clicked.connect(lambda: self.mayavi_widget.visualization.add_value_to_point(update))
            return self.button

        adjustments_layout.addWidget(create_adjust_button(self, '-5', [-5,0,0]), 1, 2) #create x-5 button
        adjustments_layout.addWidget(create_adjust_button(self, '-1', [-1,0,0]), 1, 3) #create x-1 button
        adjustments_layout.addWidget(create_adjust_button(self, '+1', [1,0,0]), 1, 4) #create x+1 button
        adjustments_layout.addWidget(create_adjust_button(self, '+5', [5,0,0]), 1, 5) #create x+5 button

        adjustments_layout.addWidget(create_adjust_button(self, '-5', [0,-5,0]), 2, 2) #create y-5 button
        adjustments_layout.addWidget(create_adjust_button(self, '-1', [0,-1,0]), 2, 3) #create y-1 button
        adjustments_layout.addWidget(create_adjust_button(self, '+1', [0,1,0]), 2, 4) #create y+1 button
        adjustments_layout.addWidget(create_adjust_button(self, '+5', [0,5,0]), 2, 5) #create y+5 button

        adjustments_layout.addWidget(create_adjust_button(self, '-5', [0,0,-5]), 3, 2) #create z-5 button
        adjustments_layout.addWidget(create_adjust_button(self, '-1', [0,0,-1]), 3, 3) #create z-1 button
        adjustments_layout.addWidget(create_adjust_button(self, '+1', [0,0,1]), 3, 4) #create z+1 button
        adjustments_layout.addWidget(create_adjust_button(self, '+5', [0,0,5]), 3, 5) #create z+5 button

        adjustments_layout.setColumnMinimumWidth(6, 45)
        adjustments_layout.setRowMinimumHeight(4, 25)

        annotation_layout.addLayout(adjustments_layout, Qt.AlignCenter)


        # VISUALIZATION OPTIONS LAYOUT
        create_header('Visualization options', visualization_layout)

        visualization_buttons_layout = QGridLayout()

        self.volume_button = QPushButton('show\nimage')
        self.volume_button.clicked.connect(self.mayavi_widget.visualization.toggle_volume)
        visualization_buttons_layout.addWidget(self.volume_button, 0, 0)

        self.trajectory_button = QPushButton("show\ntrajectory")
        self.trajectory_button.clicked.connect(lambda: self.mayavi_widget.visualization.toggle_trajectory())
        visualization_buttons_layout.addWidget(self.trajectory_button, 0, 1)

        self.all_trajectories_button = QPushButton("show all\ntrajectories")
        self.all_trajectories_button.clicked.connect(lambda: self.mayavi_widget.visualization.toggle_all_trajectories())
        visualization_buttons_layout.addWidget(self.all_trajectories_button, 0, 2)

        visualization_layout.addLayout(visualization_buttons_layout)


        # Sliders layout
        sliders_layout = QGridLayout()

        # transparency slider
        self.transparency_label = QLabel('transparency')
        self.transparency_label.setMinimumWidth(80)
        sliders_layout.addWidget(self.transparency_label, 0, 0)

        self.transparency_slider = QSlider(Qt.Horizontal)
        self.transparency_slider.setValue(5)
        self.transparency_slider.setMinimum(1.0)
        self.transparency_slider.setMaximum(20.0)
        self.transparency_slider.setSingleStep(0.1)
        self.transparency_slider.setMinimumWidth(170)
        self.transparency_slider.sliderReleased.connect(self.change_transparancy)  #past pas aan bij loslaten, kan ook bij bewegen maar is lag

        sliders_layout.addWidget(self.transparency_slider, 0, 1)

        # sphere size slider
        self.sphere_size_label = QLabel('sphere size')
        self.sphere_size_label.setMinimumWidth(80)
        sliders_layout.addWidget(self.sphere_size_label, 1, 0)

        self.sphere_size_slider = QSlider(Qt.Horizontal)
        self.sphere_size_slider.setValue(10)
        self.sphere_size_slider.setMinimum(1.0)
        self.sphere_size_slider.setMaximum(20.0)
        self.sphere_size_slider.setSingleStep(0.1)
        self.sphere_size_slider.setMinimumWidth(170)
        self.sphere_size_slider.sliderReleased.connect(self.change_sphere_size)

        sliders_layout.addWidget(self.sphere_size_slider, 1, 1)


        visualization_layout.addLayout(sliders_layout)


        # Add layouts to options page
        options_page_layout.addStretch(0)
        options_page_layout.addLayout(annotation_layout)
        options_page_layout.addLayout(visualization_layout)
        options_page_layout.addStretch(10)


    # CELL SELECTION PAGE
        cell_page_layout = QGridLayout()
        cell_page.setLayout(cell_page_layout)
        #TODO


    # Add pages to tab wiget
        tab.addTab(options_page, 'Options')
        tab.addTab(cell_page, 'Cell Selection')


    # FRAME NAVIGATION LAYOUT
        navigation_layout = QGridLayout()

        goto_button = QPushButton('go to frame')
        goto_button.clicked.connect(self.goto_frame)
        navigation_layout.addWidget(goto_button, 0, 1)

        previous_button = QPushButton('<')
        previous_button.clicked.connect(self.change_volume_model_previous)
        navigation_layout.addWidget(previous_button, 0, 2)

        next_button = QPushButton('>')
        next_button.clicked.connect(self.change_volume_model_next)
        navigation_layout.addWidget(next_button, 0, 3)

        self.frame_label = QLabel('frame 1/' + str(self.mayavi_widget.visualization.amount_of_frames-1)) #number of current frame modified when switching
        navigation_layout.addWidget(self.frame_label, 1, 3, Qt.AlignRight | Qt.AlignTop)

        
    # Add widgets to main layout
        main_layout.addWidget(tab)
        main_layout.addLayout(navigation_layout)


    # TOOLBAR, STATUSBAR, MENU
        self.setup_bar_actions()


    #popup layout #voor export
        self.popup=QDialog(self)
        self.popup.setWindowTitle("Export information")

        self.popup_layout = QGridLayout()
        x_axis_label = QLabel('X axis size')
        y_axis_label = QLabel('Y axis size')
        z_axis_label = QLabel('Z axis size')

        self.x_input = QSpinBox(maximum=10000,value=self.mayavi_widget.visualization.x_lenght)
        self.y_input = QSpinBox(maximum=10000,value=self.mayavi_widget.visualization.y_lenght)
        self.z_input = QSpinBox(maximum=10000,value=self.mayavi_widget.visualization.z_lenght)

        accept_button = QPushButton('export')
        accept_button.clicked.connect(self.popup.accept)

        cancel_button = QPushButton('cancel')
        cancel_button.clicked.connect(self.popup.reject)

        self.popup_layout.addWidget(x_axis_label,0,0)
        self.popup_layout.addWidget(self.x_input,0,1)
        self.popup_layout.addWidget(y_axis_label,1,0)
        self.popup_layout.addWidget(self.y_input,1,1)
        self.popup_layout.addWidget(z_axis_label,2,0)
        self.popup_layout.addWidget(self.z_input,2,1)
        self.popup_layout.addWidget(accept_button,3,0)
        self.popup_layout.addWidget(cancel_button,3,1)

        self.popup.setLayout(self.popup_layout)

    def setup_bar_actions(self): #beschrijf balk bovenaan met 'file'
        self.statusBar()

        loadAnnotAction = QAction(QIcon(get_filled_pixmap('graphics/load.png')), 'Load trajectories', self)
        loadAnnotAction.setShortcut(QKeySequence.Open) # Ctrl+O
        loadAnnotAction.setStatusTip('Load file containing trajectories')
        loadAnnotAction.triggered.connect(self.load_annot_dialog)

        saveAnnotAction = QAction(QIcon(get_filled_pixmap('graphics/save.png')), 'Save trajectories', self)
        saveAnnotAction.setShortcut(QKeySequence.Save) # Ctrl+S
        saveAnnotAction.setStatusTip('Save trajectories in a .xlsx file')
        saveAnnotAction.triggered.connect(self.save_annots_dialog)

        exportAction = QAction(QIcon(get_filled_pixmap('graphics/render.png')), 'Export dataset', self)
        exportAction.setShortcut('Ctrl+E')
        exportAction.setStatusTip('Export source and trajectories as dataset directory')
        exportAction.triggered.connect(self.export_dialog)

        exitAction = QAction(QIcon(get_filled_pixmap('graphics/delete.png')), 'Exit', self)
        exitAction.setShortcut(QKeySequence.Quit) # Ctrl+Q
        exitAction.setStatusTip('Exit application')
        exitAction.triggered.connect(self.close)

        #'goto' knop
        gotoAction = QAction('goto', self)
        gotoAction.setShortcut(QKeySequence.Find)
        gotoAction.setStatusTip('Go to specific frame')
        gotoAction.triggered.connect(self.goto_frame)

        #'<' en '>' knoppen
        ChangeVolumeNextAction = QAction('>', self)
        ChangeVolumeNextAction.setShortcut(QKeySequence.Find)
        ChangeVolumeNextAction.setStatusTip('Go to another volume rendering')
        ChangeVolumeNextAction.triggered.connect(self.change_volume_model_next)

        ChangeVolumePreviusAction = QAction('<', self)
        ChangeVolumePreviusAction.setShortcut(QKeySequence.Find)
        ChangeVolumePreviusAction.setStatusTip('Go to another volume rendering')
        ChangeVolumePreviusAction.triggered.connect(self.change_volume_model_previous)

    # HIDDEN HOTKEY ACTIONS
        slideLeftAction = QAction('Left', self)
        slideLeftAction.setShortcut('Left')
        slideLeftAction.setStatusTip('Slide left')
        slideLeftAction.triggered.connect(self.slide_left)

        slideRightAction = QAction('Right', self)
        slideRightAction.setShortcut('Right')
        slideRightAction.setStatusTip('Slide right')
        slideRightAction.triggered.connect(self.slide_right)

        renderAction = QAction('Render', self) #uit annot3D
        renderAction.setShortcut('R')
        renderAction.setStatusTip('Update annotation render')
        renderAction.triggered.connect(self.render)

        self.addAction(slideLeftAction)
        self.addAction(slideRightAction)
        self.addAction(renderAction)
        
    # adding menubar actions 
        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&File')
        fileMenu.addAction(loadAnnotAction)
        fileMenu.addAction(saveAnnotAction)
        fileMenu.addAction(exportAction)
        fileMenu.addAction(exitAction)

    #wat alle knoppen doen
    #dialoog windows voor save en export
    def save_annots_dialog(self):
        fname, _ = QFileDialog.getSaveFileName(self, 'Save trajectories', '.',"*.xlsx")
        #global annot3D
        if fname:
            window.mayavi_widget.visualization.save_trajectory(fname)  

    def export_dialog(self):
        outcome = self.popup.exec() # use .exec_() for Python 3.10.10
        if outcome==0: #don't do rest of function if cancel button has been pressed
            return
        x_size = self.x_input.value()
        y_size = self.y_input.value()
        z_size = self.z_input.value()
        fname, _ = QFileDialog.getSaveFileName(self, 'Save dataset directory', '.',"*.xlsx")
        if fname:
            window.mayavi_widget.visualization.export_trajectory(fname,x_size,y_size,z_size)

    def update_frame_number(self): #used to change frame number display
        frame_number = self.mayavi_widget.visualization.current_frame_number
        text = "frame "+str(frame_number)+"/"+str(self.mayavi_widget.visualization.amount_of_frames-1)
        self.frame_label.setText(text)

    def goto_frame(self): #'goto' knop popup
        # global p, annot3D

        cs, ok = QInputDialog.getText(self, "Go to frame", "Go to frame")
        if ok and cs.isnumeric(): # current frame cs must be a number
            cs = int(cs)
            if cs < 0: # frame out of range
                return
        window.mayavi_widget.visualization.goto_frame(cs)
        self.update_frame_number()

    def render(self):
        self.mayavi_widget.update_annot()

    def change_volume_model_next(self):
        window.mayavi_widget.visualization.next_frame()
        self.update_frame_number()

    def change_volume_model_previous(self):
        window.mayavi_widget.visualization.previous_frame()
        self.update_frame_number()

    def slide_left(self):
        self.change_volume_model_previous()

    def slide_right(self):
        self.change_volume_model_next()

    def change_transparancy(self):
        new_transparancy = self.transparency_slider.value()/10
        window.mayavi_widget.visualization.transparancy = new_transparancy
        window.mayavi_widget.visualization.update_volume()

    def change_sphere_size(self):
        new_size = self.sphere_size_slider.value()
        window.mayavi_widget.visualization.sphere_size = new_size
        window.mayavi_widget.visualization.redraw_all_points()


if __name__ == "__main__":
    # open new instance of app if it is not running yet
    if not QApplication.instance():
        app = QApplication(sys.argv)
    else:
        app = QApplication.instance()

    # set app style
    app.setStyle('Fusion')
    # palette = QPalette()
    # palette.setColor(QPalette.Window, QColor(53, 53, 53))
    # palette.setColor(QPalette.WindowText, Qt.white)
    # palette.setColor(QPalette.Base, QColor(25, 25, 25))
    # palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
    # palette.setColor(QPalette.ToolTipBase, Qt.black)
    # palette.setColor(QPalette.ToolTipText, Qt.white)
    # palette.setColor(QPalette.Text, Qt.white)
    # palette.setColor(QPalette.Button, QColor(53, 53, 53))
    # palette.setColor(QPalette.ButtonText, Qt.white)
    # palette.setColor(QPalette.BrightText, Qt.red)
    # palette.setColor(QPalette.Link, QColor(42, 130, 218))
    # palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
    # palette.setColor(QPalette.HighlightedText, Qt.black)
    # app.setPalette(palette)

    # create main window and start app
    window = MainWindow()
    window.show()
    sys.exit(app.exec()) # use .exec_() for Python 3.10.10
